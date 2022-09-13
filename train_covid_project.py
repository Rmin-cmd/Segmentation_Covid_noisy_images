import time
from glob import glob
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import torch
from torch.utils.data import DataLoader
import torchmetrics as tmc
import albumentations as albu
from monai.losses.dice import DiceLoss
from sklearn import metrics
import torch.nn as nn
from TEST_DATA import DriveDataset
from UTILS_TEST import epoch_time
from tqdm import tqdm
import segmentation_models_pytorch as segmp
import imageio
from torch.optim import lr_scheduler


def evaluation_metric(conf_mat):
    eps = 1e-7
    ''' for percision '''
    percision_class0 = (conf_mat[0, 0] / (np.sum(conf_mat[:, 0])+eps))
    percision_class1 = (conf_mat[1, 1] / (np.sum(conf_mat[:, 1])+eps))
    percision_class2 = (conf_mat[2, 2] / (np.sum(conf_mat[:, 2])+eps))
    percision_overall = (percision_class0+percision_class1+percision_class2)/3.0

    ''' recall '''
    recall_class0 = (conf_mat[0, 0] / np.sum(conf_mat[0, :]))
    recall_class1 = (conf_mat[1, 1] / np.sum(conf_mat[1, :]))
    recall_class2 = (conf_mat[2, 2] / np.sum(conf_mat[2, :]))
    recall_overall = (recall_class0+recall_class1+recall_class2)/3.0

    ''' F1 score '''
    f1score = 2*((percision_overall*recall_overall)/(percision_overall+recall_overall+eps))

    return percision_overall, recall_overall, f1score


def to_tensor(x,**kwargs):
    return x.transpose(2, 0, 1).astype('float64')


def to_tensor_mask(x,**kwargs):
    return x.transpose(2, 0, 1).astype('float64')


def get_preprocessing(preprocessing_fn):

    _transform = [
        albu.Lambda(image=preprocessing_fn),
        albu.Lambda(image=to_tensor, mask=to_tensor)
    ]
    return albu.Compose(_transform)


def training_augmentation():
    train_transform = [
        albu.HorizontalFlip(p=0.2),
        albu.VerticalFlip(p=0.2)
    ]
    return albu.Compose(train_transform)


def train(model, loader, optimizer, loss_fn, device):
    epoch_loss = 0.0
    model.train()
    loop = tqdm(loader)
    for idx, (x, y) in enumerate(loop):
        x = x.to(device, dtype=torch.float32)
        y = y.to(device, dtype=torch.float32)
        with torch.cuda.amp.autocast():
            y_pred = model(x)

            loss = loss_fn(y_pred, y)

        preds = (y_pred > 0.9).float()
        predicted_image = np.argmax(preds.cpu(), axis=1)
        original_image = np.argmax(y.cpu(), axis=1)
        accuracy = acc(predicted_image,original_image)
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        epoch_loss += loss.item()

    epoch_loss = epoch_loss / len(loader)
    epoch_accuracy = acc.compute()
    acc.reset()
    return epoch_loss, epoch_accuracy


def test(model, loader, loss_fn, device, conf_mat):
    epoch_loss = 0.0
    model.eval()
    i = 1
    conf_mat_p = np.zeros([3, 3])
    loop = tqdm(loader)
    for idx, (x, y) in enumerate(loop):
        x = x.to(device, dtype=torch.float32)
        y = y.to(device, dtype=torch.float32)

        predlist = torch.zeros(0, dtype=torch.long, device='cpu')
        lbllist = torch.zeros(0, dtype=torch.long, device='cpu')

        y_pred = model(x)
        loss = loss_fn(y_pred, y)
        preds = (y_pred > 0.9).float()

        predicted_image = np.argmax(preds.cpu(), axis=1)
        pred_list_test = torch.cat([predlist, predicted_image.view(-1).cpu()])

        original_image = np.argmax(y.cpu(), axis=1)
        label_list_test = torch.cat([lbllist, original_image.view(-1).cpu()])

        accuracy = acc(original_image, predicted_image)
        IOU = iou(predicted_image, original_image)

        conf_mat = metrics.confusion_matrix(label_list_test.numpy(), pred_list_test.numpy())
        if conf_mat.shape != (3, 3):
            conf_mat = np.pad(conf_mat, (0, 1), mode='constant')

        conf_mat_p += conf_mat
        predicted_image = (predicted_image.detach().numpy() * 127.5).astype(np.uint8)
        original_image = (original_image.detach().numpy() * 127.5).astype(np.uint8)

        imageio.imwrite("D:\deep learning project\Test_outputs\preds-" + str(i) + ".jpg", predicted_image[0, :, :])
        imageio.imwrite("D:\deep learning project\Test_outputs\The_main_mask-" + str(i) + ".jpg",
                        original_image[0, :, :])

        loss = loss.to(torch.float16)
        epoch_loss += loss.item()
        i = i + 1

    conf_mat_out = (conf_mat_p / (len(label_list_test.numpy()))) * 100
    epoch_loss = epoch_loss / len(loader)
    epoch_accuracy = acc.compute()
    epoch_iou = iou.compute()
    acc.reset()
    iou.reset()
    return epoch_loss,epoch_accuracy,epoch_iou,conf_mat_p,conf_mat_out


def evaluate(model, loader, loss_fn, device, conf_mat):
    epoch_loss = 0.0
    model.eval()
    Fscore = 0.0
    conf_mat_p = np.zeros([3,3])
    with torch.no_grad():
        i = 1
        loop = tqdm(loader)
        for idx, (x, y) in enumerate(loop):
            x = x.to(device, dtype=torch.float32)
            y = y.to(device, dtype=torch.float32)

            predlist = torch.zeros(0, dtype=torch.long, device='cpu')
            lbllist = torch.zeros(0, dtype=torch.long, device='cpu')

            y_pred = model(x)
            loss = loss_fn(y_pred, y)
            preds = (y_pred > 0.9).float()

            predicted_image = np.argmax(preds.cpu(), axis=1)
            pred_list_train = torch.cat([predlist, predicted_image.view(-1).cpu()])

            original_image = np.argmax(y.cpu(), axis=1)
            label_list = torch.cat([lbllist, original_image.view(-1).cpu()])

            Fscore = F_score(predicted_image,original_image)
            accuracy = acc(predicted_image,original_image)
            Precision = precision(predicted_image,original_image)
            Recall = recall(predicted_image,original_image)
            IOU = iou(predicted_image,original_image)

            conf_mat = metrics.confusion_matrix(label_list.numpy(), pred_list_train.numpy())
            if conf_mat.shape != (3,3):
                conf_mat = np.pad(conf_mat,(0,1),mode='constant')
            conf_mat_p += conf_mat

            predicted_image = (predicted_image.detach().numpy() * 127.5).astype(np.uint8)
            original_image = (original_image.detach().numpy()*127.5).astype(np.uint8)

            imageio.imwrite("D:\deep learning project\output recurrent Residual Unet\preds-" + str(i) + ".jpg",
                            predicted_image[0, :, :])
            imageio.imwrite("D:\deep learning project\output recurrent Residual Unet\The_main_mask-" + str(i) + ".jpg",
                            original_image[0, :, :])

            loss = loss.to(torch.float16)
            epoch_loss += loss.item()
            i = i + 1

        conf_mat2 = (conf_mat_p / (len(label_list.numpy())*(idx+1))) * 100
        epoch_loss = epoch_loss / len(loader)

        epoch_accuracy = acc.compute()
        epoch_fscore = F_score.compute()
        epoch_precision = precision.compute()
        epoch_recall = recall.compute()
        epoch_iou = iou.compute()

        acc.reset()
        F_score.reset()
        precision.reset()
        recall.reset()
        iou.reset()

    return epoch_loss, epoch_accuracy, conf_mat_p, conf_mat2, epoch_fscore, epoch_precision, epoch_recall, epoch_iou


if __name__ == "__main__":
    """ Seeding """
    # seeding(42)
    """ Directories """
    # create_dir("files")
    path_wb_trainloss = 'D:\deep learning project\Final results folds\Train_loss_folds.xlsx'
    path_wb_validloss = 'D:\deep learning project\Final results folds\Valid_loss_folds.xlsx'
    path_wb_validiou = 'D:\deep learning project\Final results folds\Valid_iou_folds.xlsx'
    path_wb_validfscore = 'D:\deep learning project\Final results folds\Valid_fscore_folds.xlsx'

    wb_trainloss = load_workbook(path_wb_trainloss)
    wb_validloss = load_workbook(path_wb_validloss)
    wb_validiou = load_workbook(path_wb_validiou)
    wb_validfscore = load_workbook(path_wb_validfscore)

    """ Load dataset """
    train_x = sorted(
        glob("D:\Covid noisy dataset\low_noise_stand_8\ct_2-20220807T073906Z-001\Train\jpg data\*"))
    train_y = sorted(
        glob("D:\Covid noisy dataset\low_noise_stand_8\ct_2-20220807T073906Z-001\Train\jpg mask\*"))

    valid_x = sorted(
        glob("D:\Covid noisy dataset\low_noise_stand_8\ct_2-20220807T073906Z-001\Valid\jpg data\*"))
    valid_y = sorted(
        glob("D:\Covid noisy dataset\low_noise_stand_8\ct_2-20220807T073906Z-001\Valid\jpg mask\*"))

    test_x = sorted(
        glob('D:\Covid noisy dataset\high_noise_40\ct_2-20220807T094357Z-002\Test\jpg data\*')
    )
    test_y = sorted(
        glob('D:\Covid noisy dataset\high_noise_40\ct_2-20220807T094357Z-002\Test\jpg mask\*')
    )

    data_str = f"Dataset Size:\nTrain: {len(train_x)} - Valid: {len(valid_x)}\n"
    print(data_str)

    """ Hyperparameters """
    H = 256
    W = 256
    size = (H, W)
    batch_size = 8
    LOAD_MODEL = False
    num_epochs = 50
    lr = 5e-4
    Test_flag = True
    checkpoint_path = "D:\deep learning project\Results\Imp\Pretrained_Unet++_regnetx_004_4folds_6_low noise_stand.pth.tar"
    Backbone = 'timm-regnetx_004'
    Encoder_weights = 'imagenet'
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")  ## RTX 3060 6GB
    train_loss_history = []
    valid_fscore_history = []
    valid_loss_history = []
    valid_iou_history = []
    conf_mat_valid = np.zeros([3, 3])
    conf_mat_test = np.zeros([3, 3])
    """Pre-processing for backbones"""
    model = segmp.UnetPlusPlus(encoder_name=Backbone, classes=3, encoder_weights=Encoder_weights).to(device)

    preprocessing_fn = segmp.encoders.get_preprocessing_fn(Backbone, Encoder_weights)
    """ Dataset and loader """
    train_dataset = DriveDataset(train_x, train_y, preprocessing=get_preprocessing(preprocessing_fn),
                                 augmentation=training_augmentation(), test_flag=False)
    valid_dataset = DriveDataset(valid_x, valid_y, preprocessing=get_preprocessing(preprocessing_fn), test_flag=False)
    test_dataset = DriveDataset(test_x, test_y, preprocessing=get_preprocessing(preprocessing_fn), test_flag=True)

    train_loader = DataLoader(
        dataset=train_dataset,
        batch_size=batch_size,
        shuffle=True,
        num_workers=4,
        pin_memory=True
    )

    valid_loader = DataLoader(
        dataset=valid_dataset,
        batch_size=batch_size,
        shuffle=False,
        num_workers=4,
        pin_memory=True
    )

    test_loader = DataLoader(
        dataset=test_dataset,
        shuffle=False,
        num_workers=2,
        pin_memory=True
    )

    DEVICE = 'cuda'
    #loss_fn = DiceLoss(sigmoid=True).to(device)
    class_proportion = torch.tensor([0.89893, 0.09254, 0.00852])
    class_weights = 0.89893/class_proportion
    loss_fn = nn.CrossEntropyLoss(weight=class_weights).to(device)

    optimizer = torch.optim.Adam(params=model.parameters(), lr=lr)
    #optimizer = torch.optim.Adadelta(params=model.parameters(), lr=lr)
    schedular = lr_scheduler.StepLR(optimizer, step_size=10, gamma=0.1)
    iou = tmc.classification.IoU(num_classes=3)
    F_score = tmc.classification.F1Score(num_classes=3, average="weighted", mdmc_average="samplewise")
    precision = tmc.classification.Precision(num_classes=3, mdmc_average="samplewise")
    recall = tmc.classification.Recall(num_classes=3, mdmc_average="samplewise")
    acc = tmc.classification.Accuracy(num_classes=3, mdmc_average="samplewise")

    if Test_flag is False:

        """ Training the model """
        best_iou = 0
        best_valid = float("inf")

        for epoch in range(num_epochs):

            start_time = time.time()
            train_loss, train_accuracy = train(model, train_loader, optimizer, loss_fn, device)
            valid_loss, valid_accuracy, conf_mat, conf_mat2, F1Score_valid, percision_valid, recall_valid, iou_valid = \
                evaluate(model, valid_loader, loss_fn, device, conf_mat_valid)
            np.savetxt("D:\deep learning project\confusion_matrix_Train.csv", conf_mat2, delimiter=",")

            percision_valid2, recall_valid2, F1_score2 = evaluation_metric(conf_mat)

            """ Saving the model """

            if best_iou < iou_valid and valid_loss < best_valid:
                print(f"Valid iou improved from {best_iou:2.4f} to {iou_valid:2.4f}.")

                best_iou = iou_valid
                torch.save(model.state_dict(), checkpoint_path)



            end_time = time.time()
            epoch_mins, epoch_secs = epoch_time(start_time, end_time)

            loss_train = train_loss
            acc_train = train_accuracy
            loss_valid = valid_loss
            acc_valid = valid_accuracy

            data_str = f'Epoch: {epoch + 1:02} | Epoch Time: {epoch_mins}m {epoch_secs}s\n'
            data_str += f'\tTrain Loss: {loss_train:.3f}\n'
            data_str += f'\t Train acc:{acc_train:.3f}\n'
            train_loss_history.append(loss_train)
            data_str += f'\t Val. Loss: {loss_valid:.3f}\n'
            valid_iou_history.append(iou_valid.item())
            valid_loss_history.append(loss_valid)
            data_str += f'\t Val. acc:{acc_valid:.3f}\n'
            data_str += f'\t percision2:{percision_valid2:.3f}\n'
            data_str += f'\t recall:{recall_valid2:.3f}\n'
            data_str += f'\t F1score:{F1_score2:.3f}\n'
            data_str += f'\t IOU:{iou_valid:.3f}\n'
            valid_fscore_history.append(F1_score2.item())
            print(data_str)
            '''
            if epoch > 0 and ((epoch+1) % 20 == 0):
                print('Learning rate decreased!')
                optimizer.param_groups[0]['lr'] = optimizer.param_groups[0]['lr'] / 10
            '''

        # scheduler.step(valid_loss)
        ws_trainloss = wb_trainloss.active
        ws_trainloss.append(train_loss_history)
        wb_trainloss.save(path_wb_trainloss)

        ws_validloss = wb_validloss.active
        ws_validloss.append(valid_loss_history)
        wb_validloss.save(path_wb_validloss)

        ws_validiou = wb_validiou.active
        ws_validiou.append(valid_iou_history)
        wb_validiou.save(path_wb_validiou)

        ws_validfscore = wb_validfscore.active
        ws_validfscore.append(valid_fscore_history)
        wb_validfscore.save(path_wb_validfscore)

        plt.title('train and valid loss ')
        plt.plot(train_loss_history, 'ro--', valid_loss_history, 'bo--')
        plt.legend(["train_loss", "valid_loss"])
        plt.show()
        plt.title('IOU')
        plt.plot(valid_iou_history, 'bo--')
        plt.legend(["train_acc", "valid_acc"])
        plt.show()
    else:
        model.load_state_dict(torch.load(checkpoint_path))
        test_loss, test_accuracy, test_iou, conf_mat, conf_mat2_test = test(model, test_loader, loss_fn, device,
                                                                            conf_mat_test)
        np.savetxt("D:\deep learning project\confusion_matrix_Test.csv", conf_mat2_test, delimiter=",")
        percision_test, recall_test, F1_score_test = evaluation_metric(conf_mat)

        data_str += f'\t Test. Loss: {test_loss:.3f}\n'
        data_str += f'\t IOU Loss: {test_iou:.3f}\n'
        data_str += f'\t Test. acc:{test_accuracy:.3f}\n'
        data_str += f'\t F1score:{F1_score_test:.3f}\n'
        data_str += f'\t percision:{percision_test:.3f}\n'
        data_str += f'\t recall:{recall_test:.3f}\n'
        print(data_str)
